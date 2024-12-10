import sys
import sqlite3
import openpyxl
import datetime
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QVBoxLayout, QLineEdit,
                             QWidget, QMessageBox,QListView, QAbstractItemView, QDialog,
                             QFormLayout, QDateEdit, QLabel, QToolTip,QCompleter,QFileDialog)
from PyQt5.QtCore import QStringListModel, Qt, QDate
from PyQt5.QtGui import QCursor
from datetime import datetime
from datetime import timedelta

class LoginWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Авторизация")
        self.setGeometry(100, 100, 400, 150)
        self.center_window()
        # Входные поля для имени пользователя и пароля
        self.username_input = QLineEdit(self)
        username_label = QLabel("Имя пользователя:", self)
        self.username_input.setPlaceholderText("Введите имя пользователя")
        self.password_input = QLineEdit(self)
        password_label = QLabel("Пароль:", self)
        self.password_input.setPlaceholderText("Введите пароль")
        self.password_input.setEchoMode(QLineEdit.Password)

        # Кнопка входа
        self.login_button = QPushButton("Войти", self)
        self.login_button.clicked.connect(self.login)

        # Layout для окна авторизации
        layout = QVBoxLayout()
        layout.addWidget(username_label)
        layout.addWidget(self.username_input)
        layout.addWidget(password_label)
        layout.addWidget(self.password_input)
        layout.addWidget(self.login_button)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def center_window(self):
        screen_geometry = QApplication.desktop().screenGeometry()
        window_geometry = self.frameGeometry()
        self.move((screen_geometry.width() - window_geometry.width()) // 2,
                  (screen_geometry.height() - window_geometry.height()) // 2)

    def login(self):
        username = self.username_input.text()
        password = self.password_input.text()
        # Проверка на пустые поля
        if not username.strip() or not password.strip():
            QMessageBox.warning(self, "Ошибка", "Имя пользователя и пароль не могут быть пустыми.")
            return
        try:
            conn = sqlite3.connect('library.db')
            cursor = conn.cursor()
            query = 'SELECT id, role FROM Users WHERE username = ? AND password = ?'
            cursor.execute(query, (username, password))
            user = cursor.fetchone()
            if user:
                user_id, role = user
                self.open_main_app(user_id, role)
            else:
                QMessageBox.warning(self, "Ошибка", "Неверное имя пользователя или пароль.")
            conn.close()
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка подключения к базе данных: {e}")

    def open_main_app(self, user_id, role):
        self.main_app_window = MainAppWindow(user_id, role)
        self.main_app_window.show()
        self.close()

class DatabaseManager:
    def __init__(self, db_file):
        self.db_file = db_file  # Путь к базе данных
        self.connection = None  # Соединение с базой данных

    def connect(self):
        #Подключение к базе данных.
        if not self.connection:
            self.connection = sqlite3.connect(self.db_file)

    def execute_query(self, query, params=None, fetch_all=False, fetch_one=False):
        #Выполнение запроса к базе данных.
        self.connect()  # Подключаемся к базе данных
        cursor = self.connection.cursor()
        if params:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        if fetch_all:
            return cursor.fetchall()
        elif fetch_one:
            return cursor.fetchone()
        else:
            self.connection.commit()

    def execute_non_query(self, query, params=None):
        #Выполнение SQL-запроса, который не возвращает результат (например, INSERT, UPDATE, DELETE).
        self.connect()  # Подключаемся к базе данных
        cursor = self.connection.cursor()
        if params:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        self.connection.commit()  # Подтверждаем изменения
        return cursor.rowcount  # Возвращаем количество затронутых строк

    def commit(self):
        #Метод для коммита транзакции.
        self.connection.commit()

    def close(self):
        #Закрытие соединения с базой данных.
        if self.connection:
            self.connection.close()
            self.connection = None

class MainAppWindow(QMainWindow):
    def __init__(self, user_id, role):
        super().__init__()
        self.process_expired_reservations('library.db')
        self.user_id = user_id
        self.user_role = role
        self.setWindowTitle("Библиотека")
        self.setGeometry(100, 100, 600, 400)
        self.center_window(self)
        self.user_id = user_id
        self.role = role
        self.init_ui()

    def center_window(self, window):
        screen_geometry = QApplication.desktop().screenGeometry()
        window_geometry = window.frameGeometry()
        window.move((screen_geometry.width() - window_geometry.width()) // 2,
                    (screen_geometry.height() - window_geometry.height()) // 2)

    def init_ui(self):
        layout = QVBoxLayout()
        if self.role == "reader":
            self.init_reader_ui(layout)
        else:
            self.init_librarian_ui(layout)
        self.logout_button = QPushButton("Выйти")
        self.logout_button.clicked.connect(self.logout)
        layout.addWidget(self.logout_button)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

        self.load_all_books()

    def init_reader_ui(self, layout):
        self.search_books_input = QLineEdit(self)
        self.search_books_input.setPlaceholderText("Поиск книг")
        layout.addWidget(self.search_books_input)

        self.books_list_view = QListView(self)
        self.books_list_view.setSelectionMode(QAbstractItemView.SingleSelection)
        self.books_list_view.clicked.connect(self.on_book_selected)  # Обработчик выбора книги
        layout.addWidget(self.books_list_view)

        self.reserve_button = QPushButton("Забронировать книгу")
        self.reserve_button.clicked.connect(self.reserve_book)
        layout.addWidget(self.reserve_button)

        self.view_reserved_button = QPushButton("Мои брони")
        self.view_reserved_button.clicked.connect(self.view_reserved_books)
        layout.addWidget(self.view_reserved_button)

        self.view_issued_button = QPushButton("Мои выданные книги")
        self.view_issued_button.clicked.connect(self.view_issued_books)
        layout.addWidget(self.view_issued_button)
        self.search_books_input.textChanged.connect(self.search_books)
        self.selected_book_title = None

    def init_librarian_ui(self, layout):
        self.search_books_input = QLineEdit(self)
        self.search_books_input.setPlaceholderText("Поиск книг")
        layout.addWidget(self.search_books_input)

        self.books_list_view = QListView(self)
        self.books_list_view.setSelectionMode(QAbstractItemView.SingleSelection)
        layout.addWidget(self.books_list_view)

        self.view_reservations_button = QPushButton("Просмотр броней")
        self.view_reservations_button.clicked.connect(self.view_all_reservations)
        layout.addWidget(self.view_reservations_button)

        self.issue_button = QPushButton("Выдать книгу")
        self.issue_button.clicked.connect(self.manual_issue_book)
        layout.addWidget(self.issue_button)

        self.view_issued_button = QPushButton("Просмотр выданных книг")
        self.view_issued_button.clicked.connect(self.view_all_issued_books)
        layout.addWidget(self.view_issued_button)

        # Кнопка "Посещения студентов" (ранее принадлежала администратору)
        self.view_visits_button = QPushButton("Посещения студентов")
        self.view_visits_button.clicked.connect(self.view_visits)
        layout.addWidget(self.view_visits_button)
        self.search_books_input.textChanged.connect(self.search_books)

    def on_book_selected(self, index):
        # Получаем выбранную книгу и сохраняем ее
        self.selected_book_title = index.data()

    def load_all_books(self):
        """Загружает все книги из базы данных и отображает их."""
        query = "SELECT title, author, genre FROM Books"
        try:
            db = DatabaseManager('library.db')
            books = db.execute_query(query, fetch_all=True)
            # Подготовка списка для отображения
            self.book_data = [
                {"title": book[0], "author": book[1], "genre": book[2]}
                for book in books
            ]
            book_list = [f"{book['title']} - {book['author']} ({book['genre']})" for book in self.book_data]
            model = QStringListModel(book_list)
            self.books_list_view.setModel(model)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить книги: {e}")

    def search_books(self):
        search_text = self.search_books_input.text()
        query = """
            SELECT title, author, genre, description 
            FROM Books 
            WHERE title LIKE ? OR author LIKE ? OR genre LIKE ?
        """
        try:
            # Создаем объект DatabaseManager с указанием пути к базе данных
            db = DatabaseManager('library.db')
            books = db.execute_query(query, (f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"),
                                     fetch_all=True)
            if not books:
                QMessageBox.information(self, "Результаты поиска", "Нет книг, соответствующих запросу.")
                return
            # Подготовка данных для отображения
            self.book_data = [
                {"title": book[0], "author": book[1], "genre": book[2], "description": book[3]}
                for book in books
            ]
            book_list = [f"{book['title']} - {book['author']} ({book['genre']})" for book in self.book_data]
            model = QStringListModel(book_list)
            self.books_list_view.setModel(model)
            # Настройка взаимодействия с мышью для отображения подсказок
            self.books_list_view.setMouseTracking(True)
            self.books_list_view.entered.connect(self.show_book_tooltip)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка при поиске книг: {e}")

    def show_book_tooltip(self, index):
        if index.isValid():
            book_info = self.book_data[index.row()]
            QToolTip.showText(QCursor.pos(), book_info['description'], self.books_list_view)

    def reserve_book(self):
        if not self.selected_book_title:
            QMessageBox.warning(self, "Ошибка", "Выберите книгу для бронирования.")
            return

        message_box = QMessageBox(self)
        message_box.setWindowTitle("Подтверждение бронирования")
        message_box.setText(
            "Внимание! Бронирование книги действительно в течение 24 часов. "
            "После истечения этого времени бронь автоматически снимается. "
            "Если бронь была отменена раньше срока, пожалуйста, обратитесь к библиотекарю.\n\nПродолжить?"
        )
        message_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        message_box.button(QMessageBox.Yes).setText("Да")
        message_box.button(QMessageBox.No).setText("Нет")

        reply = message_box.exec()

        if reply == QMessageBox.No:
            return

        # Получаем только название книги, убирая "Название - Автор"
        selected_book_title = self.selected_book_title.split(' - ')[0]
        db = DatabaseManager('library.db')
        try:
            # Проверяем наличие книги
            book = db.execute_query("SELECT id, status FROM Books WHERE title = ?", (selected_book_title,),
                                    fetch_one=True)
            if not book:
                QMessageBox.warning(self, "Ошибка", "Книга не найдена.")
                return
            book_id, book_status = book
            # Проверяем статус книги
            if book_status == 'booked':
                QMessageBox.warning(self, "Ошибка", "Эта книга уже забронирована.")
                return
            elif book_status == 'issued':
                QMessageBox.warning(self, "Ошибка", "Эта книга уже выдана и не может быть забронирована.")
                return
            # Проверяем, есть ли активная бронь на эту книгу
            existing_reservation = db.execute_query("""
                SELECT id FROM Reservations 
                WHERE book_id = ? AND status = 'available'
            """, (book_id,), fetch_one=True)
            if existing_reservation:
                QMessageBox.warning(self, "Ошибка", "Эта книга уже забронирована другим читателем.")
                return
            # Проверяем, не бронировал ли пользователь эту книгу ранее
            user_existing_reservation = db.execute_query("""
                SELECT id FROM Reservations 
                WHERE book_id = ? AND user_id = ? AND status = 'available'
            """, (book_id, self.user_id), fetch_one=True)
            if user_existing_reservation:
                QMessageBox.warning(self, "Ошибка", "Вы уже забронировали эту книгу.")
                return
            # Получаем текущее время с компьютера
            current_time = datetime.now()
            expiration_time = current_time + timedelta(days=1)  # +1 день для окончания брони
            # Преобразуем дату в строковый формат
            expiration_date_str = expiration_time.strftime('%H:%M %d.%m.%Y')
            # Добавляем новую запись о бронировании
            db.execute_non_query("""
                INSERT INTO Reservations (user_id, book_id, timestamp, reserve_date, status) 
                VALUES (?, ?, ?, ?, 'booked')
            """, (self.user_id, book_id, current_time.strftime('%H:%M %d.%m.%Y'), expiration_date_str))
            # Обновляем статус книги
            db.execute_non_query("""
                UPDATE Books 
                SET status = 'booked' 
                WHERE id = ?
            """, (book_id,))
            QMessageBox.information(self, "Успех", "Книга успешно забронирована!")
        except sqlite3.Error as e:
            print(f"Ошибка при бронировании книги: {e}")
            QMessageBox.warning(self, "Ошибка", f"Ошибка при бронировании книги: {e}")

    def process_expired_reservations(self, db_path):
        # Проверяет устаревшие брони (старше 24 часов) при запуске программы.
        connection = sqlite3.connect(db_path)
        cursor = connection.cursor()
        try:
            now = datetime.now()
            cursor.execute("""
                SELECT id, book_id, reserve_date, status
                FROM Reservations;
            """)
            reservations = cursor.fetchall()
            expired_reservations = []
            for reservation_id, book_id, reserve_date, status in reservations:
                try:
                    # Преобразуем строку с датой окончания брони в объект datetime
                    expiration_time = datetime.strptime(reserve_date, "%H:%M %d.%m.%Y")

                    # Если бронь просрочена
                    if status == 'booked' and now > expiration_time:
                        expired_reservations.append((reservation_id, book_id))
                except ValueError:
                    continue

            if expired_reservations:
                for reservation_id, book_id in expired_reservations:
                    # Удаляем устаревшую бронь из базы данных
                    cursor.execute("DELETE FROM Reservations WHERE id = ?", (reservation_id,))
                    # Обновляем статус книги на 'available'
                    cursor.execute("UPDATE Books SET status = 'available' WHERE id = ?", (book_id,))
                connection.commit()
        except sqlite3.Error as e:
            print(f"Ошибка при проверке устаревших броней: {e}")
        finally:
            connection.close()

    def view_reserved_books(self):
        # Отображает окно "Мои брони" с загруженными данными о бронях.
        db = DatabaseManager('library.db')
        try:
            # Проверка наличия броней у пользователя
            query = """
                SELECT Books.title, Books.author, Books.genre, Reservations.reserve_date 
                FROM Reservations
                JOIN Books ON Reservations.book_id = Books.id
                WHERE Reservations.user_id = ?
            """
            # Использование execute_query для получения данных
            reservations = db.execute_query(query, (self.user_id,), fetch_all=True)
            if not reservations:
                QMessageBox.information(self, "Мои брони", "У вас нет забронированных книг.")
                return
            # Создание окна только если есть брони
            self.reservations_window = QMainWindow()
            self.reservations_window.setWindowTitle("Мои брони")
            self.reservations_window.setGeometry(100, 100, 600, 400)
            layout = QVBoxLayout()
            # Список для отображения броней
            self.reservations_list_view = QListView()
            self.reservations_list_view.setSelectionMode(QAbstractItemView.MultiSelection)
            # Формирование списка броней
            reservation_list = [
                f"{r[0]} - {r[1]} ({r[2]}) | Забронировано до: {r[3]}"
                for r in reservations
            ]
            model = QStringListModel(reservation_list)
            self.reservations_list_view.setModel(model)
            layout.addWidget(self.reservations_list_view)

            # Кнопка для отмены брони
            self.cancel_reservation_button = QPushButton("Отменить бронь")
            self.cancel_reservation_button.clicked.connect(self.cancel_reservations)
            layout.addWidget(self.cancel_reservation_button)

            # Кнопка "Назад"
            back_button = QPushButton("Назад")
            back_button.clicked.connect(self.reservations_window.close)
            layout.addWidget(back_button)

            container = QWidget()
            container.setLayout(layout)
            self.reservations_window.setCentralWidget(container)

            # Центрируем окно
            self.center_window(self.reservations_window)
            # Показываем окно
            self.reservations_window.show()

        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при получении списка броней: {e}")
        # Закрытие соединения с базой данных
        db.close()

    def view_issued_books(self):
        #Функция для отображения списка выданных книг текущего пользователя.
        db = DatabaseManager('library.db')
        try:
            # Запрос для получения списка выданных книг, включая название, автора и жанр
            query = """
                SELECT b.title, b.author, b.genre, i.issue_date, i.return_date 
                FROM IssuedBooks i
                JOIN Books b ON i.book_id = b.id
                WHERE i.user_id = ? AND i.return_date >= date('now')
            """
            # Используем execute_query для выполнения запроса
            issued_books = db.execute_query(query, (self.user_id,), fetch_all=True)
            if issued_books:
                # Формируем строку с информацией о книгах, разделяя каждую книгу новой строкой
                books_list = "\n\n".join([
                    f"Название: {book[0]}\nАвтор: {book[1]}\nЖанр: {book[2]}\nДата выдачи: {book[3]}\nДата ожидаемого возврата: {book[4]}"
                    for book in issued_books
                ])
                QMessageBox.information(self, "Мои выданные книги", books_list)
            else:
                QMessageBox.information(self, "Мои выданные книги", "Нет выданных книг.")
        except Exception as e:
            # Обработка ошибок
            QMessageBox.warning(self, "Ошибка", f"Ошибка при просмотре выданных книг: {e}")
        # Закрытие соединения с базой данных через DatabaseManager
        db.close()

    def view_all_reservations(self):
        # Отображает окно со списком всех броней и функционалом отмены и выдачи.
        db = DatabaseManager('library.db')
        try:
            # Проверка наличия броней
            query = """
                SELECT 
                    r.id AS ReservationID,
                    u.username AS Читатель,
                    b.id AS BookID,
                    b.title AS Книга,
                    b.author AS Автор,
                    b.genre AS Жанр,
                    r.reserve_date AS Дата_бронирования
                FROM Reservations r
                JOIN Users u ON r.user_id = u.id
                JOIN Books b ON r.book_id = b.id
            """
            # Используем execute_query для выполнения запроса
            reservations = db.execute_query(query, fetch_all=True)
            if not reservations:
                QMessageBox.information(self, "Все брони", "Нет забронированных книг.")
                return
            # Создание окна только если есть брони
            self.reservations_window = QMainWindow()
            self.reservations_window.setWindowTitle("Все брони")
            self.reservations_window.setGeometry(100, 100, 600, 400)
            layout = QVBoxLayout()
            # Список для отображения броней
            self.reservations_list_view = QListView()
            self.reservations_list_view.setSelectionMode(QAbstractItemView.MultiSelection)
            # Формирование списка броней
            self.reservation_data = {
                i: {
                    "ReservationID": r[0],
                    "Username": r[1],
                    "BookID": r[2],
                    "Title": r[3],
                    "Author": r[4],
                    "Genre": r[5],
                    "ReserveDate": r[6],
                }
                for i, r in enumerate(reservations)
            }
            reservation_list = [
                f"{r['Title']} - {r['Author']} ({r['Genre']}) - Читатель: {r['Username']}, Дата: {r['ReserveDate']}"
                for r in self.reservation_data.values()
            ]
            model = QStringListModel(reservation_list)
            self.reservations_list_view.setModel(model)
            layout.addWidget(self.reservations_list_view)

            # Добавляем кнопку "Выдать книгу"
            issue_button = QPushButton("Выдать книгу по брони")
            issue_button.clicked.connect(self.issue_book_from_reservation)
            layout.addWidget(issue_button)

            # Кнопка "Отменить бронь"
            cancel_button = QPushButton("Отменить бронь")
            cancel_button.clicked.connect(self.cancel_reservations)
            layout.addWidget(cancel_button)

            # Кнопка "Назад"
            back_button = QPushButton("Назад")
            back_button.clicked.connect(self.reservations_window.close)
            layout.addWidget(back_button)

            # Устанавливаем виджет в центральное окно
            container = QWidget()
            container.setLayout(layout)
            self.reservations_window.setCentralWidget(container)
            self.center_window(self.reservations_window)
            self.reservations_window.show()
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при загрузке всех броней: {e}")
        finally:
            db.close()

    def issue_book_from_reservation(self):
        # Обработка выдачи книги из брони.
        selected_indexes = self.reservations_list_view.selectedIndexes()
        if not selected_indexes:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите книгу для выдачи.")
            return
        # Получаем выбранную бронь по индексу
        selected_index = selected_indexes[0].row()
        reservation = self.reservation_data.get(selected_index)

        if not reservation:
            QMessageBox.warning(self, "Ошибка", "Не удалось найти данные о выбранной броне.")
            return

        # Извлекаем данные из словаря reservation
        reservation_id = reservation['ReservationID']
        book_title = reservation['Title']
        book_author = reservation['Author']
        book_genre = reservation['Genre']
        user_username = reservation['Username']
        reserve_date = reservation['ReserveDate']

        # Используем DatabaseManager для работы с базой данных
        db = DatabaseManager('library.db')
        try:
            # Проверяем книгу в базе данных
            query = "SELECT id, status FROM Books WHERE title = ?"
            book = db.execute_query(query, (book_title,), fetch_one=True)
            if not book:
                QMessageBox.warning(self, "Ошибка", "Не удалось найти книгу для выдачи.")
                return
            book_id, book_status = book
            # Проверяем, забронирована ли книга
            if book_status != 'booked':
                QMessageBox.warning(self, "Ошибка", "Эта книга не забронирована и не может быть выдана.")
                return
            # Проверяем, не выдана ли уже книга
            query = "SELECT id FROM IssuedBooks WHERE book_id = ?"
            issued_book = db.execute_query(query, (book_id,), fetch_one=True)
            if issued_book:
                QMessageBox.warning(self, "Ошибка", "Эта книга уже выдана и не может быть выдана снова.")
                return

            # Окно для ввода даты выдачи и возврата
            dialog = QDialog(self)
            dialog.setWindowTitle("Выдача книги")
            layout = QFormLayout()

            # Отображаем информацию о книге
            layout.addRow("Название:", QLabel(book_title))
            layout.addRow("Автор:", QLabel(book_author))
            layout.addRow("Жанр:", QLabel(book_genre))
            layout.addRow("Пользователь:", QLabel(user_username))
            layout.addRow("Дата окончания бронирования:", QLabel(reserve_date))

            # Поле для ввода даты выдачи (по умолчанию сегодняшняя дата)
            issue_date_input = QLineEdit(datetime.now().strftime("%d.%m.%Y"))
            issue_date_input.setReadOnly(True)
            layout.addRow("Дата выдачи:", issue_date_input)

            # Поле для ввода даты возврата (по умолчанию через 2 недели)
            return_date_input = QDateEdit()
            return_date_input.setCalendarPopup(True)
            return_date_input.setDate(datetime.now().date() + timedelta(days=14))
            layout.addRow("Ожидаемая дата возврата:", return_date_input)

            # Кнопка для подтверждения выдачи
            accept_button = QPushButton("Выдать книгу")
            layout.addWidget(accept_button)
            dialog.setLayout(layout)
            # Связываем кнопку с функцией выдачи книги
            def handle_accept():
                issue_date = issue_date_input.date().toString("dd.MM.yyyy")
                return_date = return_date_input.date().toString("dd.MM.yyyy")

                try:
                    # Получаем ID пользователя, который забрал книгу
                    query = "SELECT id FROM Users WHERE username = ?"
                    user = db.execute_query(query, (user_username,), fetch_one=True)
                    if not user:
                        QMessageBox.warning(self, "Ошибка", "Не удалось найти пользователя.")
                        return
                    user_id = user[0]
                    # Добавляем запись в таблицу выданных книг
                    db.execute_query("""
                        INSERT INTO IssuedBooks (user_id, book_id, issue_date, return_date)
                        VALUES (?, ?, ?, ?)
                    """, (user_id, book_id, issue_date, return_date))
                    # Обновляем статус книги в таблице Books на "issued"
                    db.execute_query("""
                        UPDATE Books SET status = 'issued' WHERE id = ?
                    """, (book_id,))
                    # Удаляем бронь из таблицы Reservations
                    db.execute_query("DELETE FROM Reservations WHERE id = ?", (reservation_id,))
                    # Записываем дату выдачи книги и юзернейм в таблицу Visits
                    db.execute_query("""
                        INSERT INTO Visits (username, visit_date)
                        VALUES (?, ?)
                    """, (user_username, issue_date))

                    QMessageBox.information(self, "Успех", "Книга успешно выдана!")
                    self.reservations_window.close()
                    dialog.accept()
                    # Обновляем список броней
                    self.view_all_reservations()
                except Exception as e:
                    QMessageBox.critical(self, "Ошибка", f"Ошибка при выдаче книги: {e}")
            accept_button.clicked.connect(handle_accept)
            dialog.exec()
        finally:
            db.close()

    def cancel_reservations(self):
        selected_indexes = self.reservations_list_view.selectedIndexes()
        if not selected_indexes:
            QMessageBox.warning(self, "Ошибка", "Выберите хотя бы одну бронь для отмены.")
            return
        # Формируем список выбранных броней
        reservations_to_delete = [index.data() for index in selected_indexes]
        # Подтверждение удаления
        db = DatabaseManager('library.db')
        try:
            for reservation_info in reservations_to_delete:
                parts = reservation_info.split(" - ")
                if len(parts) < 2:
                    QMessageBox.warning(self, "Ошибка", "Некорректный формат строки брони.")
                    continue

                book_title = parts[0]
                author_info = parts[1].split(" (")
                if len(author_info) < 2:
                    QMessageBox.warning(self, "Ошибка", "Некорректная информация о книге.")
                    continue
                book_author = author_info[0]
                # Поиск ID брони
                query = """
                    SELECT Reservations.id, Reservations.book_id 
                    FROM Reservations 
                    JOIN Books ON Reservations.book_id = Books.id 
                    WHERE Books.title = ? AND Books.author = ?
                """
                reservation = db.execute_query(query, (book_title, book_author), fetch_one=True)

                if reservation:
                    reservation_id, book_id = reservation
                    # Удаляем бронь
                    db.execute_query("DELETE FROM Reservations WHERE id = ?", (reservation_id,))

                    # Обновляем статус книги на 'available'
                    db.execute_query("UPDATE Books SET status = 'available' WHERE id = ?", (book_id,))
                else:
                    QMessageBox.warning(self, "Ошибка", f"Бронь '{book_title}' не найдена.")

            # Применяем изменения в базе данных
            db.commit()

            QMessageBox.information(self, "Успех", "Выбранные брони успешно удалены.")
            self.reservations_window.close()

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при удалении брони: {e}")
        finally:
            db.close()
        # Обновляем список броней
        if self.role == 'reader':
            self.view_reserved_books()
        else:
            self.view_all_reservations()

    def manual_issue_book(self):
        # Окно для ручной выдачи книги с автоподстановкой названия книги и имени пользователя.
        dialog = QDialog(self)
        dialog.setWindowTitle("Выдача книги")
        layout = QFormLayout()
        # Используем DatabaseManager для работы с базой данных
        db = DatabaseManager('library.db')
        try:
            # Получение данных для автозаполнения
            books = db.execute_query("SELECT title, author, genre FROM Books WHERE status = 'available'",
                                     fetch_all=True)
            book_titles = [f"{row[0]} - {row[1]} ({row[2]})" for row in books]
            usernames = db.execute_query("SELECT username FROM Users", fetch_all=True)
            usernames = [row[0] for row in usernames]
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке данных: {e}")
            db.close()
            return
        finally:
            db.close()
        # Поле ввода названия книги с автозаполнением
        book_input = QLineEdit()
        book_completer = QCompleter(book_titles)
        book_completer.setCaseSensitivity(Qt.CaseInsensitive)
        book_input.setCompleter(book_completer)
        # Поле ввода имени пользователя с автозаполнением
        username_input = QLineEdit()
        username_completer = QCompleter(usernames)
        username_completer.setCaseSensitivity(Qt.CaseInsensitive)
        username_input.setCompleter(username_completer)
        # Поле "Дата выдачи" с сегодняшней датой (только для чтения)
        issue_date_input = QLineEdit(datetime.now().strftime("%d.%m.%Y"))
        issue_date_input.setReadOnly(True)  # Запрещаем редактирование текста пользователем
        # Поле даты возврата
        return_date_input = QDateEdit()
        return_date_input.setCalendarPopup(True)
        return_date_input.setDate(datetime.now().date() + timedelta(days=14))
        # Добавление элементов в форму
        layout.addRow("Название книги:", book_input)
        layout.addRow("Имя пользователя:", username_input)
        layout.addRow("Дата выдачи:", issue_date_input)
        layout.addRow("Дата возврата:", return_date_input)
        # Кнопка подтверждения
        submit_button = QPushButton("Выдать")
        submit_button.clicked.connect(
            lambda: self.process_manual_issue(
                dialog,
                book_input.text(),  # Передаем текст книги
                username_input.text(),  # Передаем текст имени пользователя
                issue_date_input.text(),  # Передаем дату выдачи как текст
                return_date_input.date().toString("dd.MM.yyyy")  # Передаем дату возврата в формате строки
            )
        )
        layout.addRow(submit_button)
        dialog.setLayout(layout)
        dialog.exec()

    def process_manual_issue(self, dialog, book_input, username_input, issue_date, return_date):
        # Обрабатываем выдачу книги вручную, используя базу данных
        db = DatabaseManager('library.db')
        try:
            # Получаем ID книги
            book_title, book_author = book_input.split(" - ")[0], book_input.split(" - ")[1].split(" (")[0]
            book = db.execute_query(
                "SELECT id FROM Books WHERE title = ? AND author = ? AND status = 'available'",
                (book_title, book_author), fetch_one=True
            )
            if not book:
                QMessageBox.warning(self, "Ошибка", "Книга не найдена или уже выдана.")
                dialog.reject()
                return
            book_id = book[0]
            # Получаем ID пользователя
            user = db.execute_query(
                "SELECT id FROM Users WHERE username = ?", (username_input,), fetch_one=True
            )
            if not user:
                QMessageBox.warning(self, "Ошибка", "Пользователь не найден.")
                dialog.reject()
                return
            user_id = user[0]
            # Добавляем запись в таблицу выданных книг
            db.execute_query(
                """
                INSERT INTO IssuedBooks (user_id, book_id, issue_date, return_date)
                VALUES (?, ?, ?, ?)
                """,
                (user_id, book_id, issue_date, return_date)
            )
            # Обновляем статус книги на 'issued'
            db.execute_query(
                "UPDATE Books SET status = 'issued' WHERE id = ?", (book_id,)
            )
            # Добавляем запись в таблицу посещений
            db.execute_query(
                """
                INSERT INTO Visits (username, visit_date)
                VALUES (?, ?)
                """,
                (username_input, issue_date)
            )
            QMessageBox.information(self, "Успех", "Книга успешно выдана!")
            dialog.accept()
        except Exception as e:
            QMessageBox.information(self, "Ошибка", f"Введите данные для выдачи книги.")
        finally:
            db.close()

    def view_all_issued_books(self):
        #Обновление списка выданных книг с добавлением возможности принять книги на возврат.
        db = DatabaseManager('library.db')  # Передаем путь к базе данных
        try:
            # Получаем все записи из таблицы выданных книг
            issued_books = db.execute_query("""
                SELECT IssuedBooks.id, Books.title, Books.author, Books.genre, Users.username, 
                       IssuedBooks.issue_date, IssuedBooks.return_date
                FROM IssuedBooks
                JOIN Books ON IssuedBooks.book_id = Books.id
                JOIN Users ON IssuedBooks.user_id = Users.id
                ORDER BY DATE(IssuedBooks.return_date) ASC
            """, fetch_all=True)
            # Если нет выданных книг, показываем сообщение и выходим
            if not issued_books:
                QMessageBox.information(self, "Выданные книги", "Нет выданных книг.")
                return
            # Формируем список для отображения
            book_list = []
            self.book_data = {}  # Словарь для хранения данных о книгах по индексу
            for idx, book in enumerate(issued_books):
                book_info = f"{book[1]} - {book[2]} ({book[3]}) - {book[4]} | Дата выдачи: {book[5]} | Ожидаемая дата возврата: {book[6]}"
                book_list.append(book_info)
                self.book_data[idx] = book  # Сохраняем данные книги
            # Создание окна
            self.issued_window = QMainWindow()
            self.issued_window.setWindowTitle("Выданные книги")
            self.issued_window.setGeometry(100, 100, 600, 400)  # Увеличили размер окна
            layout = QVBoxLayout()
            # Список для отображения выданных книг
            self.issued_list_view = QListView()
            self.issued_list_view.setSelectionMode(QAbstractItemView.SingleSelection)
            model = QStringListModel(book_list)
            self.issued_list_view.setModel(model)
            layout.addWidget(self.issued_list_view)
            # Кнопка "Принять книгу на возврат"
            if self.user_role == 'librarian':
                return_button = QPushButton("Принять книгу на возврат")
                return_button.clicked.connect(self.process_return_book)
                layout.addWidget(return_button)
            # Кнопка "Назад"
            back_button = QPushButton("Назад")
            back_button.clicked.connect(self.issued_window.close)
            layout.addWidget(back_button)
            # Устанавливаем виджет в центральное окно
            container = QWidget()
            container.setLayout(layout)
            self.issued_window.setCentralWidget(container)
            # Центрируем окно на экране
            self.center_window(self.issued_window)
            self.issued_window.show()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке выданных книг: {e}")

    def process_return_book(self):
        selected_indexes = self.issued_list_view.selectedIndexes()
        if not selected_indexes:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите книгу для возврата.")
            return
        selected_index = selected_indexes[0].row()
        book = self.book_data.get(selected_index)
        if not book:
            QMessageBox.warning(self, "Ошибка", "Не удалось найти данные о выбранной книге.")
            return
        issued_book_id, title, author, genre, username, issue_date, return_date = book
        # Открываем диалог для подтверждения возврата
        dialog = QDialog(self)
        dialog.setWindowTitle("Принять книгу на возврат")
        layout = QFormLayout()
        layout.addRow("Название:", QLabel(title))
        layout.addRow("Автор:", QLabel(author))
        layout.addRow("Жанр:", QLabel(genre))
        layout.addRow("Пользователь:", QLabel(username))
        layout.addRow("Дата выдачи:", QLabel(issue_date))
        layout.addRow("Ожидаемая дата возврата:", QLabel(return_date))
        # Поле для ввода фактической даты возврата
        actual_return_date_input = QDateEdit()
        actual_return_date_input.setCalendarPopup(True)
        actual_return_date_input.setDate(datetime.now().date())
        layout.addRow("Фактическая дата возврата:", actual_return_date_input)
        accept_button = QPushButton("Принять")
        layout.addWidget(accept_button)
        dialog.setLayout(layout)
        def handle_accept():
            actual_return_date = actual_return_date_input.date().toString("dd.MM.yyyy")
            # Проверка на правильность введенной даты возврата
            if actual_return_date > datetime.now().date().strftime("%d.%m.%Y"):
                QMessageBox.warning(self, "Ошибка", "Фактическая дата возврата не может быть позже сегодняшней даты.")
                return
            db = DatabaseManager('library.db')
            try:
                # Получаем book_id и user_id из таблицы IssuedBooks
                record = db.execute_query(
                    "SELECT book_id, user_id FROM IssuedBooks WHERE id = ?",
                    params=(issued_book_id,),
                    fetch_one=True
                )
                if not record:
                    QMessageBox.critical(self, "Ошибка", "Не удалось найти книгу для возврата.")
                    return
                book_id, user_id = record  # Разделяем данные записи
                # Получаем юзернейм пользователя
                user_record = db.execute_query(
                    "SELECT username FROM Users WHERE id = ?",
                    params=(user_id,),
                    fetch_one=True
                )
                if not user_record:
                    QMessageBox.warning(self, "Ошибка", "Не удалось найти пользователя.")
                    return
                user_username = user_record[0]
                # Удаляем запись о выдаче книги
                db.execute_non_query("DELETE FROM IssuedBooks WHERE id = ?", params=(issued_book_id,))
                # Обновляем статус книги на "available"
                db.execute_non_query("""
                    UPDATE Books 
                    SET status = 'available'
                    WHERE id = ?
                """, params=(book_id,))
                # Записываем возврат как посещение (дата возврата и юзернейм)
                db.execute_non_query("""
                    INSERT INTO Visits (username, visit_date)
                    VALUES (?, ?)
                """, params=(user_username, actual_return_date))
                QMessageBox.information(self, "Успех", "Книга успешно возвращена!")
                self.issued_window.close()
                dialog.accept()
                # Обновляем список выданных книг
                self.view_all_issued_books()
            except sqlite3.Error as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при возврате книги: {e}")
            finally:
                db.close()
        accept_button.clicked.connect(handle_accept)
        dialog.exec()

    def view_visits(self):
        # Экспортирует уникальные данные о посещениях студентов в файл Excel.
        db = DatabaseManager('library.db')
        try:
            # Собираем уникальные посещения (по пользователю и дате)
            visits_data = db.execute_query("""
                SELECT DISTINCT Users.full_name, Users.username, Users.student_group, 
                       Visits.visit_date
                FROM Visits
                JOIN Users ON Visits.username = Users.username
                ORDER BY Visits.visit_date
            """, fetch_all=True)
            if not visits_data:
                QMessageBox.warning(self, "Ошибка", "Нет данных о посещениях для экспорта.")
                return
            options = QFileDialog.Options()
            file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить файл", "","Excel Files (*.xlsx);;All Files (*)", options=options)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Посещения"
            headers = ["ФИО", "Username", "Группа", "Дата посещения"]
            header_font = Font(bold=True)
            ws.append(headers)
            for col_num, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_num).font = header_font
            for visit in visits_data:
                full_name, username, group, visit_date = visit
                ws.append((full_name, username, group, visit_date))
            for col_num in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0
                for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                # Устанавливаем ширину столбца
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Сохраняем файл
            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Файл успешно сохранен в: {file_path}")
        except Exception as e:
            QMessageBox.information(self, "Ошибка", f"Выберите путь для экспортирования данных.")
        finally:
            db.close()
    def logout(self):
        self.close()
        self.login_window = LoginWindow()
        self.login_window.show()
if __name__ == "__main__":
    app = QApplication(sys.argv)
    login_window = LoginWindow()
    login_window.show()
    sys.exit(app.exec_())
